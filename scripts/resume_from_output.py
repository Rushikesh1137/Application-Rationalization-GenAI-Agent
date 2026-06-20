"""Resume an interrupted ECR run from a partial Excel output.

The script keeps completed L3 rows from a previous workbook, reruns only L3s
whose rows were written as exception/human-review fallback rows, and writes a
new merged workbook.
"""

from __future__ import annotations

import os
import sys
import traceback
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from src.graph import build_graph
from src.orchestrator import (
    DEFAULT_INPUT_FILE,
    DEFAULT_INPUT_SHEET,
    DEFAULT_LOG_DIR,
    DEFAULT_OUTPUT_DIR,
    _safe_output_stem,
    error_state_for_l3,
    initial_state_for_l3,
    load_inventory,
    select_l3_groups,
    write_calibration_summary,
    write_error_log,
)
from src.output_writer import CAPABILITY_LOSS_COLUMN, OUTPUT_COLUMNS, write_excel
from src.state import ECRState


def _clean(value: object) -> str:
    return str(value or "").strip()


def _read_rows(workbook_path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(workbook_path, data_only=True)
    try:
        worksheet = workbook["ECR Results"]
        headers = [_clean(cell.value) for cell in worksheet[1]]
        rows: list[dict[str, str]] = []
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            row = {headers[index]: _clean(value) for index, value in enumerate(values)}
            if row.get("App Name"):
                rows.append(row)
        return rows
    finally:
        workbook.close()


def _failed_l3s(rows: list[dict[str, str]]) -> set[str]:
    failed: set[str] = set()
    for row in rows:
        rationale = row.get("Rationale", "").casefold()
        if "exception occurred while processing this cluster" in rationale:
            failed.add(row.get("Final L3", ""))
    return {item for item in failed if item}


def _state_from_completed_rows(l3_name: str, group, rows_by_l3: dict[str, list[dict[str, str]]]) -> ECRState:
    rows = rows_by_l3[l3_name]
    decisions = [{column: row.get(column, "") for column in OUTPUT_COLUMNS} for row in rows]
    return {
        "l3_name": l3_name,
        "apps": group.to_dict(orient="records"),
        "max_retries": 5,
        "function_tags": {row["App Name"]: row.get("Function", "") for row in rows},
        "failed_apps": [],
        "validator_feedback": {},
        "ecr_decisions": decisions,
        "capability_loss": {row["App Name"]: row.get(CAPABILITY_LOSS_COLUMN, "") for row in rows},
        "function_mismatch": None,
        "retry_count": 0,
        "status": "done",
    }


def resume_from_output(previous_output: Path) -> Path:
    load_dotenv(ROOT / ".env")
    input_file = Path(os.getenv("INPUT_FILE", DEFAULT_INPUT_FILE))
    input_sheet = os.getenv("INPUT_SHEET") or DEFAULT_INPUT_SHEET
    output_dir = Path(os.getenv("OUTPUT_DIR", DEFAULT_OUTPUT_DIR))
    log_dir = Path(os.getenv("LOG_DIR", DEFAULT_LOG_DIR))
    output_dir.mkdir(parents=True, exist_ok=True)
    log_dir.mkdir(parents=True, exist_ok=True)

    previous_rows = _read_rows(previous_output)
    failed_l3_names = _failed_l3s(previous_rows)
    rows_by_l3: dict[str, list[dict[str, str]]] = {}
    for row in previous_rows:
        rows_by_l3.setdefault(row["Final L3"], []).append(row)

    inventory = load_inventory(input_file, input_sheet)
    groups = select_l3_groups(inventory)
    graph = build_graph()
    final_states: list[ECRState] = []

    print(f"Previous output: {previous_output}")
    print(f"Failed L3s to rerun: {len(failed_l3_names)}")
    for name in failed_l3_names:
        print(f"- {name}")

    for l3_name, group in groups:
        if l3_name not in failed_l3_names and l3_name in rows_by_l3:
            final_states.append(_state_from_completed_rows(l3_name, group, rows_by_l3))
            continue

        print(f"Rerunning L3: {l3_name} ({len(group)} apps)")
        state = initial_state_for_l3(l3_name, group)
        try:
            final_state = graph.invoke(
                state,
                config={"recursion_limit": state["max_retries"] * 6 + 10},
            )
        except Exception as error:
            log_path = write_error_log(log_dir, l3_name, error)
            final_state = error_state_for_l3(l3_name, group, error)
            print(f"  Status: error; Log: {log_path}")
            traceback.print_exception(type(error), error, error.__traceback__)
        else:
            print(
                f"  Status: {final_state['status']}; "
                f"Retries: {final_state['retry_count']}; "
                f"Rows: {len(final_state.get('ecr_decisions', []))}"
            )
        final_states.append(final_state)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"{_safe_output_stem(input_file)}_ecr_results_resumed_{timestamp}.xlsx"
    write_excel(final_states, output_path)
    write_calibration_summary(log_dir, f"resumed_{timestamp}", final_states)
    print(f"Done. Resumed Excel results in: {output_path}")
    return output_path


if __name__ == "__main__":
    if len(sys.argv) != 2:
        raise SystemExit("Usage: python scripts/resume_from_output.py output/previous_results.xlsx")
    resume_from_output(Path(sys.argv[1]))
