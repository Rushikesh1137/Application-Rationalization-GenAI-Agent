from pathlib import Path
import csv
import os
import sys

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from src.agents.ecr_recommender import (
    CAPABILITY_LOSS_COLUMN,
    detect_function_mismatch,
    normalize_decisions,
    normalize_decisions_with_issues,
)
from src.graph import increment_retry_node
from src.routing import route_after_validation
from src.output_writer import OUTPUT_COLUMNS, write_excel, write_tsv
from src.validator import validator_node
from src.web_search import WebEvidence, build_web_evidence, format_web_evidence


def base_state(**updates):
    state = {
        "l3_name": "Data Integration",
        "apps": [],
        "max_retries": 5,
        "function_tags": {},
        "failed_apps": [],
        "validator_feedback": {},
        "ecr_decisions": [],
        "capability_loss": {},
        "function_mismatch": None,
        "retry_count": 0,
        "status": "tagging",
    }
    state.update(updates)
    return state



def test_web_search_disabled_by_default():
    old_enabled = os.environ.pop("ENABLE_WEB_SEARCH", None)
    try:
        evidence = build_web_evidence(
            [{"Name": "Example App", "Description": "", "Vendor": ""}],
            "Example L3",
        )
        assert evidence == {}
    finally:
        if old_enabled is not None:
            os.environ["ENABLE_WEB_SEARCH"] = old_enabled


def test_web_evidence_format_is_prompt_safe():
    text = format_web_evidence(
        {
            "Example App": [
                WebEvidence(
                    title="Example Source",
                    url="https://example.com/product",
                    snippet="Example App is a product suite module.",
                )
            ]
        }
    )
    assert "External evidence" in text
    assert "Example App" in text
    assert "example.com" in text
    assert "Do not invent a retained app outside" in text

def test_retry_increment_single_place():
    state = base_state(retry_count=2)
    update = increment_retry_node(state)
    assert update == {"retry_count": 3, "status": "tagging"}



def test_retry_exhaustion_routes_to_agent_2_when_tags_exist():
    state = base_state(
        apps=[{"Name": "App A"}, {"Name": "App B"}],
        function_tags={"App A": "Workflow Tool", "App B": "Workflow Tool"},
        failed_apps=["App A"],
        retry_count=5,
        max_retries=5,
    )
    assert route_after_validation(state) == "agent_2"


def test_retry_exhaustion_stops_when_tags_missing():
    state = base_state(
        apps=[{"Name": "App A"}, {"Name": "App B"}],
        function_tags={"App A": "Workflow Tool"},
        failed_apps=["App B"],
        retry_count=5,
        max_retries=5,
    )
    assert route_after_validation(state) == "end_with_flag"

def test_function_equal_l3_is_exact_equality_only():
    apps = [
        {"Name": "Bad", "Vendor": "", "Description": "", "L3": "Data Integration"},
        {"Name": "Good", "Vendor": "", "Description": "", "L3": "Data Integration"},
    ]
    state = base_state(
        apps=apps,
        function_tags={"Bad": "Data Integration", "Good": "ETL Platform"},
        status="validating",
    )
    result = validator_node(state)
    assert "Bad" in result["failed_apps"]
    assert "Good" not in result["failed_apps"]


def test_function_mismatch_checks_eliminate_and_consolidate():
    decisions = [
        {"App Name": "Elim A", "Recommendation": "Eliminate", "App to be Retained": "Target A"},
        {"App Name": "Con A", "Recommendation": "Consolidate", "App to be Retained": "Target B"},
    ]
    mismatch = detect_function_mismatch(
        decisions,
        {
            "Elim A": "Legacy Workflow",
            "Target A": "ETL Platform",
            "Con A": "File Transfer",
            "Target B": "Managed File Transfer",
        },
    )
    assert mismatch is not None
    assert len(mismatch["mismatches"]) == 2
    assert {item["recommendation"] for item in mismatch["mismatches"]} == {
        "Eliminate",
        "Consolidate",
    }


def test_eliminate_target_missing_flips_before_mismatch():
    apps = [
        {"Name": "App A", "Vendor": "", "Description": "Standalone app", "L3": "Test L3"},
        {"Name": "App B", "Vendor": "", "Description": "Different app", "L3": "Test L3"},
    ]
    state = base_state(
        apps=apps,
        l3_name="Test L3",
        function_tags={"App A": "Workflow Tool", "App B": "Reporting Tool"},
    )
    normalized = normalize_decisions(
        state,
        [
            {
                "App Name": "App A",
                "Final L3": "Test L3",
                "Function": "Workflow Tool",
                "Recommendation": "Eliminate",
                "Rationale": "An outside app should replace this app.",
                "App to be Retained": "External Target",
                CAPABILITY_LOSS_COLUMN: "",
            },
            {
                "App Name": "App B",
                "Final L3": "Test L3",
                "Function": "Reporting Tool",
                "Recommendation": "Retain",
                "Rationale": "This app remains active.",
                "App to be Retained": "App B",
                CAPABILITY_LOSS_COLUMN: "",
            },
        ],
    )
    assert normalized[0]["Recommendation"] == "Retain"
    assert normalized[0]["App to be Retained"] == "App A"
    assert detect_function_mismatch(normalized, state["function_tags"]) is None


def test_legacy_modern_pair_can_eliminate_with_blank_capability_loss():
    apps = [
        {
            "Name": "Legacy ETL",
            "Vendor": "",
            "Description": "Legacy deprecated ETL platform scheduled for retirement.",
            "L3": "Data Integration",
        },
        {
            "Name": "Modern ETL",
            "Vendor": "",
            "Description": "Strategic modern ETL platform for the same workflows.",
            "L3": "Data Integration",
        },
    ]
    state = base_state(
        apps=apps,
        function_tags={"Legacy ETL": "ETL Platform", "Modern ETL": "ETL Platform"},
    )
    normalized = normalize_decisions(
        state,
        [
            {
                "App Name": "Legacy ETL",
                "Final L3": "Data Integration",
                "Function": "ETL Platform",
                "Recommendation": "Eliminate",
                "Rationale": "Modern ETL covers the same workflows and is the strategic platform.",
                "App to be Retained": "Modern ETL",
                CAPABILITY_LOSS_COLUMN: "",
            },
            {
                "App Name": "Modern ETL",
                "Final L3": "Data Integration",
                "Function": "ETL Platform",
                "Recommendation": "Retain",
                "Rationale": "This is the strategic platform for the capability.",
                "App to be Retained": "Modern ETL",
                CAPABILITY_LOSS_COLUMN: "Not applicable",
            },
        ],
    )
    assert normalized[0]["Recommendation"] == "Eliminate"
    assert normalized[0]["App to be Retained"] == "Modern ETL"
    assert normalized[0][CAPABILITY_LOSS_COLUMN] == ""
    assert normalized[1][CAPABILITY_LOSS_COLUMN] == ""
    assert detect_function_mismatch(normalized, state["function_tags"]) is None


def test_eliminate_target_outside_cluster_flips_to_retain():
    apps = [
        {"Name": "Legacy ETL", "Vendor": "", "Description": "Legacy deprecated ETL app.", "L3": "Data Integration"},
    ]
    state = base_state(apps=apps, function_tags={"Legacy ETL": "ETL Platform"})
    normalized, issues = normalize_decisions_with_issues(
        state,
        [
            {
                "App Name": "Legacy ETL",
                "Final L3": "Data Integration",
                "Function": "ETL Platform",
                "Recommendation": "Eliminate",
                "Rationale": "A platform outside the cluster should replace it.",
                "App to be Retained": "External ETL",
                CAPABILITY_LOSS_COLUMN: "",
            }
        ],
    )
    assert normalized[0]["Recommendation"] == "Retain"
    assert normalized[0][CAPABILITY_LOSS_COLUMN] == ""
    assert issues[0]["action"] == "auto_corrected_to_retain"


def test_eliminate_target_same_app_flips_to_retain():
    apps = [
        {"Name": "Legacy ETL", "Vendor": "", "Description": "Legacy deprecated ETL app.", "L3": "Data Integration"},
    ]
    state = base_state(apps=apps, function_tags={"Legacy ETL": "ETL Platform"})
    normalized, issues = normalize_decisions_with_issues(
        state,
        [
            {
                "App Name": "Legacy ETL",
                "Final L3": "Data Integration",
                "Function": "ETL Platform",
                "Recommendation": "Eliminate",
                "Rationale": "The app should be eliminated.",
                "App to be Retained": "Legacy ETL",
                CAPABILITY_LOSS_COLUMN: "",
            }
        ],
    )
    assert normalized[0]["Recommendation"] == "Retain"
    assert "itself" in normalized[0]["Rationale"]
    assert issues[0]["action"] == "auto_corrected_to_retain"


def test_consolidate_can_have_blank_capability_loss_without_review():
    apps = [
        {"Name": "Duplicate App", "Vendor": "duplicate", "Description": "Same product", "L3": "Test L3"},
        {"Name": "Parent App", "Vendor": "duplicate", "Description": "Same product", "L3": "Test L3"},
    ]
    state = base_state(
        l3_name="Test L3",
        apps=apps,
        function_tags={"Duplicate App": "Workflow Tool", "Parent App": "Workflow Tool"},
    )
    normalized, issues = normalize_decisions_with_issues(
        state,
        [
            {
                "App Name": "Duplicate App",
                "Final L3": "Test L3",
                "Function": "Workflow Tool",
                "Recommendation": "Consolidate",
                "Rationale": "This is a duplicate CMDB entry.",
                "App to be Retained": "Parent App",
            },
            {
                "App Name": "Parent App",
                "Final L3": "Test L3",
                "Function": "Workflow Tool",
                "Recommendation": "Retain",
                "Rationale": "This app remains active.",
                "App to be Retained": "Parent App",
                CAPABILITY_LOSS_COLUMN: "",
            },
        ],
    )
    assert normalized[0]["Recommendation"] == "Consolidate"
    assert normalized[0][CAPABILITY_LOSS_COLUMN] == ""
    assert issues == []



def test_consolidation_anchor_is_marked_consolidate():
    apps = [
        {"Name": "Main CDM", "Vendor": "", "Description": "Primary CDM platform", "L3": "Accounts Receivable"},
        {"Name": "CDM Mortgage", "Vendor": "", "Description": "Mortgage instance of CDM", "L3": "Accounts Receivable"},
    ]
    state = base_state(
        l3_name="Accounts Receivable",
        apps=apps,
        function_tags={
            "Main CDM": "Payment exception review",
            "CDM Mortgage": "Payment exception review",
        },
    )
    normalized, issues = normalize_decisions_with_issues(
        state,
        [
            {
                "App Name": "Main CDM",
                "Final L3": "Accounts Receivable",
                "Function": "Payment exception review",
                "Recommendation": "Retain",
                "Rationale": "This is the primary CDM platform.",
                "App to be Retained": "Main CDM",
                CAPABILITY_LOSS_COLUMN: "",
            },
            {
                "App Name": "CDM Mortgage",
                "Final L3": "Accounts Receivable",
                "Function": "Payment exception review",
                "Recommendation": "Consolidate",
                "Rationale": "This instance rolls into the main CDM platform.",
                "App to be Retained": "Main CDM",
                CAPABILITY_LOSS_COLUMN: "",
            },
        ],
    )
    from src.agents.ecr_recommender import mark_consolidation_anchors

    marked = mark_consolidation_anchors(normalized)
    assert marked[0]["Recommendation"] == "Consolidate"
    assert marked[0]["App to be Retained"] == "Main CDM"
    assert marked[1]["Recommendation"] == "Consolidate"
    assert issues == []

def test_error_state_writes_review_rows_and_excel():
    state = base_state(
        l3_name="Broken L3",
        apps=[{"Name": "Broken App"}],
        status="error",
        error_type="RuntimeError",
        error_message="boom",
    )
    temp_dir = Path("output/test_hardening_tmp")
    temp_dir.mkdir(parents=True, exist_ok=True)
    tsv_path = temp_dir / "out.tsv"
    xlsx_path = temp_dir / "direct.xlsx"
    try:
        write_tsv([state], tsv_path)
        rows = list(csv.DictReader(tsv_path.open(encoding="utf-8"), delimiter="\t"))
        assert rows[0]["Recommendation"] == "Retain"
        assert rows[0]["Function"] == "Needs Review"
        assert "human review" in rows[0]["Rationale"].casefold()
        assert CAPABILITY_LOSS_COLUMN in rows[0]

        write_excel([state], xlsx_path)
        assert xlsx_path.exists()
        workbook = load_workbook(xlsx_path, read_only=True)
        try:
            assert workbook.sheetnames[0] == "Summary"
            assert "ECR Results" in workbook.sheetnames
            assert workbook["ECR Results"].max_column == len(OUTPUT_COLUMNS)
        finally:
            workbook.close()
    finally:
        if tsv_path.exists():
            tsv_path.unlink()
        if xlsx_path.exists():
            xlsx_path.unlink()
        if temp_dir.exists():
            temp_dir.rmdir()


def run_all():
    test_web_search_disabled_by_default()
    test_web_evidence_format_is_prompt_safe()
    test_retry_increment_single_place()
    test_retry_exhaustion_routes_to_agent_2_when_tags_exist()
    test_retry_exhaustion_stops_when_tags_missing()
    test_function_equal_l3_is_exact_equality_only()
    test_function_mismatch_checks_eliminate_and_consolidate()
    test_eliminate_target_missing_flips_before_mismatch()
    test_legacy_modern_pair_can_eliminate_with_blank_capability_loss()
    test_eliminate_target_outside_cluster_flips_to_retain()
    test_eliminate_target_same_app_flips_to_retain()
    test_consolidate_can_have_blank_capability_loss_without_review()
    test_consolidation_anchor_is_marked_consolidate()
    test_error_state_writes_review_rows_and_excel()
    print("hardening tests passed")


if __name__ == "__main__":
    run_all()


