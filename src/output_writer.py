import csv
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.state import ECRState


CAPABILITY_LOSS_COLUMN = "Capability Loss if Eliminated"
OUTPUT_COLUMNS = [
    "App Name",
    "Final L3",
    "Function",
    "Recommendation",
    "Rationale",
    "App to be Retained",
    CAPABILITY_LOSS_COLUMN,
]

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
ELIMINATE_FILL = PatternFill(fill_type="solid", fgColor="F4CCCC")
SUMMARY_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")


def _clean(value: object) -> str:
    return str(value or "").strip()


def _review_rationale_for_state(state: ECRState) -> str:
    if state["status"] == "error":
        error_type = _clean(state.get("error_type")) or "ProcessingError"
        return (
            f"An exception occurred while processing this cluster: {error_type}. "
            "This app needs human review."
        )
    return (
        "Retry budget was exhausted before a clean recommendation was produced. "
        "This app needs human review."
    )


def _fallback_rows_for_state(state: ECRState) -> list[dict[str, str]]:
    """Create review rows when a cluster does not have clean ECR decisions."""
    rows: list[dict[str, str]] = []
    function_tags = state.get("function_tags", {})
    capability_loss = state.get("capability_loss", {})
    rationale = _review_rationale_for_state(state)
    for app in state["apps"]:
        app_name = _clean(app.get("Name"))
        rows.append(
            {
                "App Name": app_name,
                "Final L3": state["l3_name"],
                "Function": _clean(function_tags.get(app_name)) or "Needs Review",
                "Recommendation": "Retain",
                "Rationale": rationale,
                "App to be Retained": app_name,
                CAPABILITY_LOSS_COLUMN: _clean(capability_loss.get(app_name)),
            }
        )
    return rows


def rows_for_state(state: ECRState) -> list[dict[str, str]]:
    """Return normalized output rows for one final cluster state."""
    rows = state.get("ecr_decisions") or _fallback_rows_for_state(state)
    capability_loss = state.get("capability_loss", {})
    normalized_rows: list[dict[str, str]] = []
    for row in rows:
        normalized = {column: _clean(row.get(column)) for column in OUTPUT_COLUMNS}
        app_name = normalized["App Name"]
        if not normalized[CAPABILITY_LOSS_COLUMN]:
            normalized[CAPABILITY_LOSS_COLUMN] = _clean(capability_loss.get(app_name))
        if state["status"] == "max_retries_hit" and normalized["Rationale"]:
            normalized["Rationale"] = f"{normalized['Rationale']} Needs human review."
        elif state["status"] == "error" and "human review" not in normalized["Rationale"].casefold():
            normalized["Rationale"] = f"{normalized['Rationale']} Needs human review."
        normalized_rows.append(normalized)
    return normalized_rows


def rows_for_states(states: list[ECRState]) -> list[dict[str, str]]:
    """Return normalized output rows for all final cluster states."""
    rows: list[dict[str, str]] = []
    for state in states:
        rows.extend(rows_for_state(state))
    return rows


def _format_results_worksheet(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    recommendation_column = OUTPUT_COLUMNS.index("Recommendation") + 1
    capability_column = OUTPUT_COLUMNS.index(CAPABILITY_LOSS_COLUMN) + 1

    for row in ws.iter_rows(min_row=2):
        recommendation = _clean(row[recommendation_column - 1].value)
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        capability_text = _clean(row[capability_column - 1].value)
        if recommendation == "Eliminate" and capability_text:
            row[capability_column - 1].fill = ELIMINATE_FILL

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        if column_letter == "E":
            ws.column_dimensions[column_letter].width = 85
        elif column_letter == "G":
            ws.column_dimensions[column_letter].width = 60
        else:
            ws.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 38)


def _append_section_header(ws, title: str) -> None:
    ws.append([title])
    row_number = ws.max_row
    cell = ws.cell(row=row_number, column=1)
    cell.font = Font(bold=True)
    cell.fill = SUMMARY_FILL


def _capability_issue_rows(states: list[ECRState]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for state in states:
        for issue in state.get("capability_loss_issues", []):
            rows.append(
                {
                    "L3": _clean(issue.get("l3_name")) or state["l3_name"],
                    "App Name": _clean(issue.get("app_name")),
                    "Issue": _clean(issue.get("issue")),
                    "Action": _clean(issue.get("action")),
                }
            )
    return rows


def _write_summary_sheet(wb: Workbook, states: list[ECRState], rows: list[dict[str, str]]) -> None:
    ws = wb.active
    ws.title = "Summary"

    recommendation_counts = Counter(row["Recommendation"] for row in rows)
    l3_counts = Counter(row["Final L3"] for row in rows)
    max_retry_clusters = [state["l3_name"] for state in states if state["status"] == "max_retries_hit"]
    capability_issues = _capability_issue_rows(states)

    ws.append(["Run timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["Total apps processed", len(rows)])
    ws.append([])

    _append_section_header(ws, "Counts by Recommendation")
    ws.append(["Recommendation", "Count"])
    for recommendation in ["Retain", "Eliminate", "Consolidate"]:
        ws.append([recommendation, recommendation_counts.get(recommendation, 0)])
    ws.append([])

    _append_section_header(ws, "Counts by L3")
    ws.append(["L3", "Count"])
    for l3_name, count in sorted(l3_counts.items()):
        ws.append([l3_name, count])
    ws.append([])

    _append_section_header(ws, "Clusters That Hit Max Retries")
    ws.append(["L3"])
    if max_retry_clusters:
        for l3_name in max_retry_clusters:
            ws.append([l3_name])
    else:
        ws.append(["None"])
    ws.append([])

    _append_section_header(ws, "Capability Loss Warnings Or Auto-Corrected Rows")
    ws.append(["L3", "App Name", "Issue", "Action"])
    if capability_issues:
        for issue in capability_issues:
            ws.append([issue["L3"], issue["App Name"], issue["Issue"], issue["Action"]])
    else:
        ws.append(["None", "", "", ""])

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 26


def write_excel(states: list[ECRState], output_path: Path) -> Path:
    """Write final ECR rows directly to an Excel workbook."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    rows = rows_for_states(states)

    wb = Workbook()
    _write_summary_sheet(wb, states, rows)
    ws = wb.create_sheet("ECR Results")
    ws.append(OUTPUT_COLUMNS)

    for row in rows:
        ws.append([row[column] for column in OUTPUT_COLUMNS])

    _format_results_worksheet(ws)
    wb.save(output_path)

    check = load_workbook(output_path, read_only=True)
    try:
        worksheet = check["ECR Results"]
        if worksheet.max_column != len(OUTPUT_COLUMNS):
            raise ValueError(
                f"Excel output has {worksheet.max_column} columns, expected {len(OUTPUT_COLUMNS)}."
            )
        if "Summary" not in check.sheetnames:
            raise ValueError("Excel output is missing the Summary sheet.")
    finally:
        check.close()

    return output_path


def write_tsv(states: list[ECRState], output_path: Path) -> Path:
    """Legacy helper: write final ECR rows to TSV."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=OUTPUT_COLUMNS, delimiter="\t")
        writer.writeheader()
        for row in rows_for_states(states):
            writer.writerow(row)
    return output_path


def write_excel_from_tsv(tsv_path: Path) -> Path:
    """Legacy helper: create an Excel workbook beside a TSV output file."""
    xlsx_path = tsv_path.with_suffix(".xlsx")
    with tsv_path.open("r", encoding="utf-8", newline="") as file:
        rows = list(csv.reader(file, delimiter="\t"))

    wb = Workbook()
    ws = wb.active
    ws.title = "ECR Results"

    for row in rows:
        ws.append(row)

    _format_results_worksheet(ws)
    wb.save(xlsx_path)
    return xlsx_path


